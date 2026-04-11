import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

folder_path = "D:\\a"

all_rows = []
seen_item_nos = set()  # ✅ Theo dõi itemNo đã thêm

for filename in os.listdir(folder_path):
    if filename.endswith(".json"):
        filepath = os.path.join(folder_path, filename)
        category_name = filename.replace(".json", "")

        with open(filepath, encoding='utf-8') as f:
            data = json.load(f)

        if not data:
            continue

        for item in data:
            item_no = item.get('itemNo', '')

            # ✅ Bỏ qua nếu itemNo đã tồn tại
            if item_no and item_no in seen_item_nos:
                continue
            if item_no:
                seen_item_nos.add(item_no)

            seo = item.get('seoName', '')
            url = f"https://www.winmart.vn/products/{seo}" if seo else ''

            all_rows.append([
                category_name,
                item_no,
                item.get('name', ''),
                item.get('shortDescription', ''),
                item.get('brandName', ''),
                item.get('uomName', ''),
                item.get('barcode', ''),
                item.get('price', ''),
                item.get('salePrice', ''),
                item.get('quantity', ''),
                item.get('scaleQuantity', ''),
                item.get('mch3Name', ''),
                item.get('mch4Name', ''),
                item.get('categoryName', ''),
                url
            ])

print(f"📦 Tổng sản phẩm sau dedup: {len(all_rows)}")

# 📊 Tạo Excel (giữ nguyên phần này)
wb = Workbook()
ws = wb.active
ws.title = "All Products"

headers = [
    'Danh mục (file)', 'Mã SP', 'Tên sản phẩm', 'Mô tả ngắn',
    'Thương hiệu', 'Đơn vị', 'Barcode', 'Giá gốc', 'Giá sale',
    'Tồn kho', 'Scale Qty', 'Nhóm hàng 3', 'Nhóm hàng 4', 'Category', 'Link'
]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 22

for row_data in all_rows:
    ws.append(row_data)
    last_row = ws.max_row
    url = row_data[-1]
    name_cell = ws.cell(row=last_row, column=3)
    if url and isinstance(url, str) and url.startswith("http"):
        name_cell.hyperlink = url
        name_cell.font = Font(color="0563C1", underline="single")

column_widths = {
    'A': 30, 'B': 12, 'C': 40, 'D': 35, 'E': 20,
    'F': 10, 'G': 16, 'H': 12, 'I': 12, 'J': 10,
    'K': 10, 'L': 20, 'M': 20, 'N': 20, 'O': 50,
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
wb.save("sp2.xlsx")
print(f"✅ Xong! Tổng {len(all_rows)} sản phẩm → sp.xlsx")